"""Edge-case audit — exercises corners of the API to surface bugs that the
happy-path smoke tests miss.

Categories covered:
  1. CSV parsing edge cases (encoding, delimiters, malformed cells)
  2. State machine boundaries (re-finalise guards, step rollback)
  3. Iterative re-solve edge cases (infeasibility, floor invalidation)
  4. Module 1 input boundary values (carve-out, seasonality, goal values)
  5. Module 5 / Monte Carlo edge cases (small budget, single platform)
  6. Catalog & lookup completeness for every built-in platform
"""
from __future__ import annotations

import pytest

from core.wizard_state import (
    WizardState, ALLOWED_PLATFORMS,
    GOAL_AW, GOAL_EN, GOAL_WT, GOAL_LG,
)
from core.kpi_config import KPI_CONFIG, get_kpi_rows
from core.csv_import import parse_platform_csv, SUPPORTED_PLATFORMS
from modules.module1 import (
    complete_module1_and_advance, Module1ValidationError,
)
from modules.module2 import run_module2
from modules.module3 import finalise_module3_from_inputs
from modules.module4 import run_module4
from modules.module5 import (
    run_module5, run_module5_montecarlo, build_module5_input_from_state,
    Module5ValidationError, PLATFORM_EFFECTIVE_MINIMUMS_PER_MONTH,
)
from modules.module6 import run_module6
from modules.module7 import run_module7


# ─────────────────────────────────────────────────────────────────────────────
# Group 1: CSV parsing edge cases
# ─────────────────────────────────────────────────────────────────────────────

def test_csv_empty_returns_error():
    assert "error" in parse_platform_csv(b"", "fb")


def test_csv_only_headers_returns_error():
    assert "error" in parse_platform_csv(b"Reach,Impressions\n", "fb")


def test_csv_semicolon_delimiter_is_sniffed():
    """European locale exports use ; as delimiter."""
    csv = b"Reach;Impressions;Amount Spent\n100000;200000;1000\n"
    result = parse_platform_csv(csv, "fb")
    assert "error" not in result
    assert result["budget"] == pytest.approx(1000.0)
    assert result["kpis"]["FB_AW_REACH"] == pytest.approx(100000.0)


def test_csv_with_bom_prefix():
    """Google sometimes prefixes UTF-8 BOM; should be stripped."""
    csv = "﻿Impressions,Clicks,Conversions,Cost\n1000,50,5,200\n".encode("utf-8")
    result = parse_platform_csv(csv, "go_search")
    assert "error" not in result
    assert result["kpis"].get("GO_SEARCH_LG_CONVERSIONS") == pytest.approx(5.0)


def test_csv_latin1_encoding_falls_back():
    """Some Windows exports use Latin-1 not UTF-8."""
    # 'Coût' (cost in French) — non-ASCII in Latin-1
    csv = "Impressions,Clicks,Conversions,Cost\n1000,50,5,200\n".encode("latin-1")
    result = parse_platform_csv(csv, "go_search")
    assert "error" not in result


def test_csv_with_blanks_and_dashes():
    """Cells like '--' or 'N/A' shouldn't crash; treated as missing."""
    csv = b"Reach,Impressions,Amount Spent\n--,200000,1000\nN/A,100000,500\n"
    result = parse_platform_csv(csv, "fb")
    # No reach values → no FB_AW_REACH; impressions still aggregated
    assert "FB_AW_REACH" not in result["kpis"]
    assert result["kpis"]["FB_AW_IMPRESSION"] == pytest.approx(300000.0)


def test_csv_with_currency_symbols():
    """Spend cells like '£1,200.50' should parse to 1200.50."""
    csv = b'Reach,Amount Spent\n100000,"\xc2\xa31,200.50"\n'
    result = parse_platform_csv(csv, "fb")
    assert result["budget"] == pytest.approx(1200.50)


def test_csv_google_ctr_decimal_form():
    """When Google exports CTR as decimal (0.045), parse to 0.045, not 0.00045."""
    csv = b"Impressions,Clicks,CTR,Conversions,Cost\n100000,4500,0.045,300,1000\n"
    result = parse_platform_csv(csv, "go_search")
    # CTR provided as decimal — should remain 0.045 (NOT divide by 100 again)
    assert result["kpis"]["GO_SEARCH_EN_CTR"] == pytest.approx(0.045)


def test_csv_thousand_separator_in_numbers():
    csv = b'Reach,Impressions,Amount Spent\n"200,000","1,500,000","3,200.50"\n'
    result = parse_platform_csv(csv, "fb")
    assert result["kpis"]["FB_AW_REACH"] == pytest.approx(200000.0)
    assert result["budget"] == pytest.approx(3200.50)


def test_csv_extra_irrelevant_columns_ignored():
    csv = b"Date,Account,Reach,Impressions,Amount Spent,Ad Set,Region\n"
    csv += b"2024-01-01,ACME,100000,200000,1000,Set 1,UK\n"
    result = parse_platform_csv(csv, "fb")
    assert result["kpis"]["FB_AW_REACH"] == pytest.approx(100000.0)


def test_csv_supported_platforms_covers_key_marketers():
    """The platforms a real marketer most needs CSV import for must be supported.
    Google is split into Search / Display / PMax — Search is the most common
    paid-media surface, so it must be on this list."""
    for p in ("fb", "ig", "li", "go_search", "go_display", "go_pmax", "tt", "yt"):
        assert p in SUPPORTED_PLATFORMS, f"{p!r} missing from CSV support"


def test_csv_google_ctr_bare_percentage_form():
    """Google exports sometimes show CTR as '4.5' (meaning 4.5%) without
    the '%' suffix.  The parser needs to recognise this — a bare 4.5
    cannot mean 'CTR = 450%' literally.  Rate KPIs must be normalised
    into [0, 1] regardless of presentation."""
    csv = b"Impressions,Clicks,CTR,Conversions,Cost\n100000,4500,4.50,300,1000\n"
    result = parse_platform_csv(csv, "go_search")
    # Should be 0.045 (CTR of 4.5%), not 4.5
    assert result["kpis"]["GO_SEARCH_EN_CTR"] == pytest.approx(0.045), (
        f"Bare '4.50' in CTR column should normalise to 0.045, got "
        f"{result['kpis']['GO_SEARCH_EN_CTR']}"
    )


def test_csv_engagement_rate_bare_percentage_form():
    """Same problem on TikTok/IG/LI: CTR/ER reported as a percentage
    without the % sign should not produce a 'rate' value > 1."""
    csv = b"Impressions,Engagement rate,Cost\n100000,5.5,1000\n"
    result = parse_platform_csv(csv, "tt")
    if "TT_EN_ENGRATERATE" in result["kpis"]:
        assert result["kpis"]["TT_EN_ENGRATERATE"] <= 1.0, (
            f"Rate KPIs must be in [0,1]; got {result['kpis']['TT_EN_ENGRATERATE']}"
        )


def test_csv_with_totals_row_doesnt_double_count():
    """Google's standard export appends a 'Total' row that contains the sum
    of all campaign rows.  If we naively sum, we double the spend."""
    csv = (
        b"Campaign,Impressions,Clicks,Conversions,Cost\n"
        b"Camp A,500000,2000,100,500\n"
        b"Camp B,500000,2000,100,500\n"
        b'"Total --",1000000,4000,200,1000\n'
    )
    result = parse_platform_csv(csv, "go_search")
    # Expected spend is 500 + 500 = 1000 (NOT 1000 + 1000 = 2000 from
    # naively summing the Total row too).
    assert result["budget"] == pytest.approx(1000.0), (
        f"Total row was summed into the budget: expected £1000, got £{result['budget']}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Group 2: Module 1 boundary values
# ─────────────────────────────────────────────────────────────────────────────

def test_module1_carveout_exact_boundary_rejected():
    s = WizardState()
    with pytest.raises(Module1ValidationError):
        complete_module1_and_advance(s, raw_objectives=["aw"], raw_budget=10000.0,
                                     raw_test_and_learn_pct=0.5)


def test_module1_carveout_just_below_boundary_accepted():
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["aw"], raw_budget=10000.0,
                                 raw_test_and_learn_pct=0.499)
    assert s.test_and_learn_pct == pytest.approx(0.499)


def test_module1_negative_carveout_rejected():
    s = WizardState()
    with pytest.raises(Module1ValidationError):
        complete_module1_and_advance(s, raw_objectives=["aw"], raw_budget=10000.0,
                                     raw_test_and_learn_pct=-0.1)


def test_module1_seasonality_at_boundaries():
    """0.1 and 10.0 should be the inclusive limits."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_seasonality_index={"lg": 0.1})
    assert s.seasonality_index["lg"] == pytest.approx(0.1)
    s2 = WizardState()
    complete_module1_and_advance(s2, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_seasonality_index={"lg": 10.0})
    assert s2.seasonality_index["lg"] == pytest.approx(10.0)


def test_module1_seasonality_just_outside_bounds_rejected():
    s = WizardState()
    with pytest.raises(Module1ValidationError):
        complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                     raw_seasonality_index={"lg": 0.099})
    s2 = WizardState()
    with pytest.raises(Module1ValidationError):
        complete_module1_and_advance(s2, raw_objectives=["lg"], raw_budget=10000.0,
                                     raw_seasonality_index={"lg": 10.001})


def test_module1_seasonality_for_unselected_goal_dropped_silently():
    """A multiplier for a goal not in valid_goals shouldn't raise; just drop."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_seasonality_index={"lg": 1.5, "aw": 0.8})
    assert "aw" not in s.seasonality_index
    assert s.seasonality_index["lg"] == pytest.approx(1.5)


def test_module1_goal_value_zero_is_dropped():
    """Zero is non-actionable for a value weight; should be filtered out."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg", "aw"], raw_budget=10000.0,
                                 raw_goal_values={"lg": 100.0, "aw": 0.0})
    assert "aw" not in s.goal_value_per_unit
    assert s.goal_value_per_unit["lg"] == pytest.approx(100.0)


def test_module1_negative_goal_value_rejected():
    s = WizardState()
    with pytest.raises(Module1ValidationError):
        complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                     raw_goal_values={"lg": -50.0})


# ─────────────────────────────────────────────────────────────────────────────
# Group 4: Module 5 LP edge cases
# ─────────────────────────────────────────────────────────────────────────────

def test_module5_single_platform_single_goal():
    """The minimal feasible LP: one platform, one goal.  Should allocate the
    entire budget to the single cell."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "lg", "priority_2": None}})
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "historical_days": 30, "kpis": {"FB_LG_LEADS": 100.0}},
    })
    run_module4(s)
    run_module5(s)
    base = s.module5_scenario_bundle.results_by_scenario["base"]
    fb_total = sum(base.budget_per_platform_goal["fb"].values())
    assert fb_total > 9000.0  # ≈ entire budget (minus tiny rounding)


def test_module5_infeasible_floors_caught():
    """If floors sum to more than the budget, base scenario must raise."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=1000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb", "li"],
                priorities_input={
                    "fb": {"priority_1": "lg", "priority_2": None},
                    "li": {"priority_1": "lg", "priority_2": None},
                })
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "historical_days": 30, "kpis": {"FB_LG_LEADS": 60.0}},
        "li": {"budget": 3000.0, "historical_days": 30, "kpis": {"LI_LG_LEADS": 40.0}},
    })
    run_module4(s)
    # Floors sum to £1500 > total £1000 → must fail
    s.min_spend_per_platform = {"fb": 800.0, "li": 700.0}
    with pytest.raises(Module5ValidationError):
        run_module5(s)


def test_module5_tiny_budget_still_works():
    """Budgets at the low end (£100) should still produce a valid plan."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=100.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "lg", "priority_2": None}})
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "historical_days": 30, "kpis": {"FB_LG_LEADS": 60.0}},
    })
    run_module4(s)
    s.min_spend_per_platform = {"fb": 0.0}
    run_module5(s)
    base = s.module5_scenario_bundle.results_by_scenario["base"]
    assert base.total_budget_used <= 100.0 + 1e-3


def test_module5_huge_budget_no_overflow():
    """Budgets at the high end (£10M) shouldn't break the LP."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10_000_000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb", "li"],
                priorities_input={
                    "fb": {"priority_1": "lg", "priority_2": None},
                    "li": {"priority_1": "lg", "priority_2": None},
                })
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 50000.0, "historical_days": 90, "kpis": {"FB_LG_LEADS": 500.0}},
        "li": {"budget": 50000.0, "historical_days": 90, "kpis": {"LI_LG_LEADS": 400.0}},
    })
    run_module4(s)
    run_module5(s)
    base = s.module5_scenario_bundle.results_by_scenario["base"]
    assert base.total_budget_used <= 10_000_000.0 + 10.0


# ─────────────────────────────────────────────────────────────────────────────
# Group 5: State machine boundaries
# ─────────────────────────────────────────────────────────────────────────────

def test_running_module5_before_module4_raises():
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "lg", "priority_2": None}})
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "historical_days": 30, "kpis": {"FB_LG_LEADS": 60.0}},
    })
    # Skip Module 4
    with pytest.raises(Exception):  # FlowStateError or similar
        run_module5(s)


def test_re_finalising_module1_raises():
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0)
    with pytest.raises(Exception):
        complete_module1_and_advance(s, raw_objectives=["aw"], raw_budget=20000.0)


# ─────────────────────────────────────────────────────────────────────────────
# Group 6: Monte Carlo edge cases
# ─────────────────────────────────────────────────────────────────────────────

def _mc_ready_state():
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "lg", "priority_2": None}})
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "historical_days": 30, "kpis": {"FB_LG_LEADS": 60.0}},
    })
    run_module4(s)
    return s


def test_montecarlo_single_platform_works():
    """One platform = nothing to shrink toward, but MC should still produce
    a valid (trivial) distribution."""
    s = _mc_ready_state()
    mc = run_module5_montecarlo(s, n_trials=20, seed=1)
    assert mc.n_trials > 0
    assert len(mc.per_platform) == 1
    # With one platform and a deterministic single cell, allocation is fixed
    assert mc.per_platform[0].cv >= 0.0


def test_montecarlo_exact_max_trials_accepted():
    s = _mc_ready_state()
    # n_trials=1000 is the documented max
    mc = run_module5_montecarlo(s, n_trials=1000, seed=1)
    assert mc.n_trials > 0


def test_montecarlo_over_max_trials_rejected():
    s = _mc_ready_state()
    with pytest.raises(Module5ValidationError):
        run_module5_montecarlo(s, n_trials=1001, seed=1)


def test_montecarlo_zero_seed_reproducible():
    """Seed=0 is a valid seed; must be reproducible like any other."""
    s1 = _mc_ready_state()
    s2 = _mc_ready_state()
    a = run_module5_montecarlo(s1, n_trials=15, seed=0)
    b = run_module5_montecarlo(s2, n_trials=15, seed=0)
    assert a.per_platform[0].mean == pytest.approx(b.per_platform[0].mean)


# ─────────────────────────────────────────────────────────────────────────────
# Group 7: Catalog completeness
# ─────────────────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("platform", sorted(ALLOWED_PLATFORMS))
def test_every_builtin_platform_has_all_four_goals(platform):
    """Every built-in platform should have at least one KPI for each of the
    four canonical goals (aw, en, wt, lg) — otherwise some Module-2
    selections would silently drop the platform from the LP."""
    for goal in (GOAL_AW, GOAL_EN, GOAL_WT, GOAL_LG):
        rows = get_kpi_rows(platform, goal)
        assert rows, f"{platform!r} has no KPI for goal {goal!r}"


@pytest.mark.parametrize("platform", sorted(ALLOWED_PLATFORMS))
def test_every_builtin_platform_has_effective_minimum(platform):
    """Effective spend warnings depend on PLATFORM_EFFECTIVE_MINIMUMS_PER_MONTH;
    a missing entry produces a silent zero threshold."""
    assert platform in PLATFORM_EFFECTIVE_MINIMUMS_PER_MONTH, (
        f"{platform!r} missing from effective-minimums table"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Group 8: Iterative re-solve edge cases
# ─────────────────────────────────────────────────────────────────────────────

def test_resolve_to_infeasible_floors_raises():
    """If the user dials a floor higher than the budget, the re-solve must
    raise rather than silently producing a wrong plan."""
    from tests.smoke_test import _run_pipeline_to_module5
    s = _run_pipeline_to_module5()
    run_module6(s)

    # Set floors that exceed the budget
    s.min_spend_per_platform = {"fb": 6000.0, "ig": 6000.0, "li": 6000.0}
    s.module4_finalised = False
    s.module5_finalised = False
    s.module6_finalised = False
    s.current_step = 4
    run_module4(s)
    with pytest.raises(Module5ValidationError):
        run_module5(s)


def test_resolve_with_lower_carveout_releases_budget():
    """A smaller carve-out should make more budget available to the LP."""
    from tests.smoke_test import _run_pipeline_to_module5
    s = _run_pipeline_to_module5()
    s.test_and_learn_pct = 0.30
    run_module2  # unused; included to keep imports linted

    # First re-solve with heavy carve-out
    s.module4_finalised = False
    s.module5_finalised = False
    s.module6_finalised = False
    s.current_step = 4
    s.test_and_learn_pct = 0.30
    run_module4(s)
    run_module5(s)
    used_heavy = s.module5_scenario_bundle.results_by_scenario["base"].total_budget_used

    # Re-solve with light carve-out
    s.module4_finalised = False
    s.module5_finalised = False
    s.module6_finalised = False
    s.current_step = 4
    s.test_and_learn_pct = 0.05
    run_module4(s)
    run_module5(s)
    used_light = s.module5_scenario_bundle.results_by_scenario["base"].total_budget_used

    assert used_light > used_heavy, (
        f"Smaller carve-out should free more budget for the LP; "
        f"heavy ({used_heavy:.0f}) vs light ({used_light:.0f})"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Group 9: Forecast edge cases
# ─────────────────────────────────────────────────────────────────────────────

def test_module6_forecast_with_zero_allocation():
    """A platform-goal cell with zero allocation must produce no forecast row,
    not a divide-by-zero or unit error."""
    from tests.smoke_test import _run_pipeline_to_module5
    s = _run_pipeline_to_module5()
    # Override allocation: zero IG entirely
    lp = s.module5_scenario_bundle.results_by_scenario["base"]
    for g in list(lp.budget_per_platform_goal.get("ig", {}).keys()):
        lp.budget_per_platform_goal["ig"][g] = 0.0
    run_module6(s)
    base_fc = s.module6_scenario_result.results_by_scenario["base"]
    ig_rows = [r for r in base_fc.rows if r.platform == "ig"]
    assert all(r.allocated_budget > 0 for r in ig_rows), (
        "IG had zero allocations across all goals; should produce no forecast rows"
    )


def test_csv_negative_spend_doesnt_corrupt_budget():
    """Refund/credit rows shouldn't drag the spend total below zero into
    nonsense — Module 3 expects budget > 1."""
    csv = b"Campaign,Reach,Amount Spent\nCamp A,100000,1500\nCredit,0,-200\n"
    result = parse_platform_csv(csv, "fb")
    # Sums to £1300; positive and meaningful
    assert result["budget"] > 1.0


def test_csv_with_only_irrelevant_columns_no_kpis_matched():
    """A CSV with columns we don't recognise should produce kpis={} and
    a missing list, not crash."""
    csv = b"Region,Population\nLondon,9000000\n"
    result = parse_platform_csv(csv, "fb")
    assert result.get("kpis", {}) == {} or all(v == 0 for v in result["kpis"].values())
    assert result.get("missing_kpis"), "Expected missing_kpis to be populated"


def test_csv_wrong_platform_doesnt_crash():
    """Upload a Google export to the FB parser: should produce mostly missing
    fields but not crash."""
    csv = b"Impressions,Clicks,CTR,Conversions,Cost\n100000,4500,4.50%,300,1000\n"
    result = parse_platform_csv(csv, "fb")
    # FB pattern won't match Google's CTR / Conversions columns; only
    # 'Impressions' (matches FB_AW_IMPRESSION) and 'Cost' (FB _budget) match.
    assert "error" not in result
    assert result["kpis"].get("FB_AW_IMPRESSION") == pytest.approx(100000.0)


def test_csv_blank_value_cells_dont_zero_aggregate():
    """A column that's mostly blank shouldn't aggregate to zero (and then get
    filtered as <=0) silently."""
    csv = (
        b"Campaign,Reach,Impressions,Amount Spent\n"
        b"Camp A,,200000,1000\n"
        b"Camp B,,300000,1500\n"
    )
    result = parse_platform_csv(csv, "fb")
    # No Reach values → FB_AW_REACH should be in missing, not silently 0
    assert "FB_AW_REACH" in result["missing_kpis"]
    assert result["kpis"].get("FB_AW_IMPRESSION") == pytest.approx(500000.0)


def test_pipeline_with_all_builtin_platforms():
    """Run the full pipeline with every built-in platform selected at once —
    stress test for the LP and the catalog lookups.  Google contributes
    three distinct surfaces (Search / Display / PMax), so this is now a
    12-platform stress test rather than 10."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["aw", "lg"], raw_budget=100000.0,
                                 raw_duration_days=30)
    builtins = [
        "fb", "ig", "li", "yt", "tt", "pt", "tw", "sn", "rd",
        "go_search", "go_display", "go_pmax",
    ]
    run_module2(s, selected_platforms=builtins,
                priorities_input={p: {"priority_1": "lg", "priority_2": "aw"}
                                  for p in builtins})
    # Each platform gets one credible LG count
    lg_vars = {
        "fb": "FB_LG_LEADS", "ig": "IG_LG_LEADS", "li": "LI_LG_LEADS",
        "yt": "YT_LG_LEADS", "tt": "TT_LG_LEADS", "pt": "PT_LG_LEADS",
        "tw": "TW_LG_LEADS", "sn": "SN_LG_LEADS", "rd": "RD_LG_LEADS",
        "go_search":  "GO_SEARCH_LG_CONVERSIONS",
        "go_display": "GO_DISPLAY_LG_CONVERSIONS",
        "go_pmax":    "GO_PMAX_LG_CONVERSIONS",
    }
    inputs = {}
    for p in builtins:
        inputs[p] = {
            "budget": 5000.0, "historical_days": 60,
            "kpis": {lg_vars[p]: 50.0 + 5 * builtins.index(p)},
        }
    finalise_module3_from_inputs(s, platform_inputs=inputs)
    run_module4(s)
    run_module5(s)
    run_module6(s)
    base = s.module5_scenario_bundle.results_by_scenario["base"]
    assert base.total_budget_used > 0
    # Every selected platform should appear in budget_per_platform
    for p in builtins:
        assert p in base.budget_per_platform


def test_montecarlo_with_high_carveout_still_valid():
    """A near-maximum carve-out (49%) should still let Monte Carlo run."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_duration_days=30, raw_test_and_learn_pct=0.49)
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "lg", "priority_2": None}})
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "historical_days": 30, "kpis": {"FB_LG_LEADS": 60.0}},
    })
    run_module4(s)
    mc = run_module5_montecarlo(s, n_trials=20, seed=1)
    assert mc.n_trials > 0


def test_composition_sums_engagement_components_excluding_clicks():
    """A Meta CSV with separate Reactions/Comments/Shares/Saves columns
    should sum them into FB_EN_ENGAGEMENT, deliberately NOT including the
    Link Clicks column (which lives in FB_WT_CLICKS)."""
    csv = (
        b"Campaign,Reach,Reactions,Comments,Shares,Saves,Link clicks,Amount spent\n"
        b"Camp A,100000,4000,1500,1500,1500,3000,2000\n"
    )
    result = parse_platform_csv(csv, "fb")
    assert "error" not in result
    # Engagement = 4000 + 1500 + 1500 + 1500 = 8500 — clicks NOT included
    assert result["kpis"]["FB_EN_ENGAGEMENT"] == pytest.approx(8500.0)
    # Clicks still extracted separately into WT
    assert result["kpis"]["FB_WT_CLICKS"] == pytest.approx(3000.0)


def test_composition_falls_back_to_bundled_with_warning():
    """A CSV that only has Meta's bundled 'Post engagement' column should
    still produce a value, but the breakdown must flag used_fallback=True
    and the rationale must warn about the double-count risk."""
    csv = (
        b"Reach,Post engagement,Link clicks,Amount spent\n"
        b"100000,12000,3000,2000\n"
    )
    result = parse_platform_csv(csv, "fb")
    # Falls back to the bundled value
    assert result["kpis"]["FB_EN_ENGAGEMENT"] == pytest.approx(12000.0)
    bd = result["kpi_breakdown"]["FB_EN_ENGAGEMENT"]
    assert bd["used_fallback"] is True
    assert "double-count" in bd["rationale"].lower() or \
           "double counting" in bd["rationale"].lower()


def test_composition_breakdown_shows_components():
    """The breakdown should expose the per-component values so the UI can
    display 'engagement = reactions(4000) + comments(1500) + ...'."""
    csv = (
        b"Reactions,Comments,Shares,Saves,Amount spent\n"
        b"4000,1500,1500,1500,2000\n"
    )
    result = parse_platform_csv(csv, "fb")
    components = result["kpi_breakdown"]["FB_EN_ENGAGEMENT"]["components"]
    assert len(components) == 4
    values = [c["value"] for c in components]
    assert values == [4000.0, 1500.0, 1500.0, 1500.0]
    # Operator should be 'sum'
    assert result["kpi_breakdown"]["FB_EN_ENGAGEMENT"]["operator"] == "sum"


def test_composition_partial_components_produce_partial_sum():
    """If only some component columns are present (e.g. Reactions + Comments
    but no Shares / Saves), the engagement sum should reflect only what
    was found — not zero, not the bundled fallback."""
    csv = b"Reactions,Comments,Amount spent\n4000,1500,2000\n"
    result = parse_platform_csv(csv, "fb")
    assert result["kpis"]["FB_EN_ENGAGEMENT"] == pytest.approx(5500.0)
    bd = result["kpi_breakdown"]["FB_EN_ENGAGEMENT"]
    assert bd["used_fallback"] is False
    assert len(bd["components"]) == 2


def test_missing_data_detector_flags_empty_platform():
    """A platform selected in M2 with no KPIs at all in M3 must be flagged
    with reason 'no_platform_data' so the results UI can surface it."""
    from modules.module5 import detect_missing_data_cells

    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb", "li"],
                priorities_input={
                    "fb": {"priority_1": "lg", "priority_2": None},
                    "li": {"priority_1": "lg", "priority_2": None},
                })
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "kpis": {}},   # ← empty kpis
        "li": {"budget": 3000.0, "kpis": {"LI_LG_LEADS": 80.0}},
    })

    issues = detect_missing_data_cells(s)
    fb_issues = [i for i in issues if i.platform == "fb"]
    assert fb_issues, "FB with empty kpis dict should be flagged"
    assert fb_issues[0].reason == "no_platform_data"
    assert fb_issues[0].goal is None


def test_missing_data_detector_flags_cell_gap():
    """A platform with some goals covered but not others should produce
    per-cell 'no_cell_data' entries for the gaps."""
    from modules.module5 import detect_missing_data_cells

    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["aw", "lg"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "aw", "priority_2": "lg"}})
    # FB has AW data but NO LG data
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "kpis": {"FB_AW_REACH": 200000.0}},
    })

    issues = detect_missing_data_cells(s)
    lg_gaps = [i for i in issues if i.platform == "fb" and i.goal == "lg"]
    assert lg_gaps, "FB·LG missing should be flagged"
    assert lg_gaps[0].reason == "no_cell_data"


def test_missing_data_detector_silent_when_complete():
    """When every prioritised cell has data, the detector should return [].
    No false positives — happy-path users shouldn't see spurious warnings."""
    from modules.module5 import detect_missing_data_cells
    from tests.smoke_test import _run_pipeline_to_module5

    s = _run_pipeline_to_module5()
    assert detect_missing_data_cells(s) == []


def test_module3_none_kpi_value_raises_clean_valueerror():
    """Passing None as a KPI value used to crash with TypeError from inside
    float().  Should now raise ValueError with a useful message."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["aw"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "aw", "priority_2": None}})
    with pytest.raises(ValueError, match="must be numeric"):
        finalise_module3_from_inputs(s, platform_inputs={
            "fb": {"budget": 3000.0, "kpis": {"FB_AW_REACH": None}},
        })


def test_module3_string_kpi_value_raises_clean_valueerror():
    """Strings that aren't numeric should also raise ValueError, not
    propagate a raw ValueError from float() without the variable name."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["aw"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "aw", "priority_2": None}})
    with pytest.raises(ValueError, match="must be numeric"):
        finalise_module3_from_inputs(s, platform_inputs={
            "fb": {"budget": 3000.0, "kpis": {"FB_AW_REACH": "not a number"}},
        })


def test_csv_template_round_trips_through_parser():
    """The template a user downloads, fills in, and re-uploads should
    parse cleanly into the same canonical KPI structure — no off-by-one
    column-name mismatches between generator and parser."""
    from core.csv_import import generate_csv_template, SUPPORTED_PLATFORMS

    for platform in SUPPORTED_PLATFORMS:
        template = generate_csv_template(platform)
        assert template, f"No template for {platform!r}"
        # The example row in the template is a valid CSV that should parse
        # without errors when fed back into the same parser.
        parsed = parse_platform_csv(template, platform)
        assert "error" not in parsed, (
            f"Template for {platform!r} doesn't round-trip: {parsed.get('error')}"
        )
        # At least the budget column matched and produced a number
        assert parsed.get("budget"), (
            f"Template for {platform!r} produced no budget value"
        )


def test_csv_template_for_meta_lists_engagement_components():
    """The FB template should expose Reactions, Comments, Shares, Saves
    as separate columns (not just one 'Engagement' bucket) so users see
    the broken-out atoms the composer wants."""
    from core.csv_import import generate_csv_template

    template = generate_csv_template("fb").decode("utf-8")
    header = template.split("\n", 1)[0].lower()
    for needle in ("reaction", "comment", "share", "save"):
        assert needle in header, (
            f"FB template missing {needle!r} column. Header: {header}"
        )


def test_csv_template_rate_kpis_show_percent_example():
    """Rate KPI columns (CTR, engagement rate) should have a '%' example
    row value so users immediately see the expected format."""
    from core.csv_import import generate_csv_template

    template = generate_csv_template("go_search").decode("utf-8")
    lines = template.split("\n")
    assert len(lines) >= 2
    header_lower = lines[0].lower()
    if "ctr" in header_lower:
        # Find CTR column index, check example row contains '%' somewhere
        # (we don't enforce that specific column has %, just that the
        # example row contains a percentage form for rate KPIs).
        assert "%" in lines[1], (
            "Go template has CTR column but no '%' example value to "
            "signal rate format"
        )


def test_money_helper_uses_current_currency_from_state():
    """The money() helper should pick up the WizardState's currency so a
    USD plan doesn't render £ everywhere."""
    import streamlit as st
    from app import money, _current_currency_symbol

    # Default (no state) → £
    if "wizard_state" in st.session_state:
        del st.session_state["wizard_state"]
    assert _current_currency_symbol() == "£"
    assert money(1234.56) == "£1,234.56"

    # USD state
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_currency="USD")
    st.session_state["wizard_state"] = s
    assert _current_currency_symbol() == "$"
    assert money(1234.56) == "$1,234.56"

    # EUR state
    s2 = WizardState()
    complete_module1_and_advance(s2, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_currency="EUR")
    st.session_state["wizard_state"] = s2
    assert money(1234.56) == "€1,234.56"

    # Explicit override still works (used by PDF export)
    assert money(1234.56, currency_symbol="$") == "$1,234.56"
    # Cleanup so subsequent tests aren't affected
    del st.session_state["wizard_state"]


def test_composition_recompose_with_user_weights():
    """Simulate the Option C override: take the parsed breakdown, apply
    per-component weights, recompose with the same operator."""
    csv = (
        b"Reactions,Comments,Shares,Saves,Amount spent\n"
        b"4000,1500,1500,1500,2000\n"
    )
    parsed = parse_platform_csv(csv, "fb")
    components = parsed["kpi_breakdown"]["FB_EN_ENGAGEMENT"]["components"]
    operator = parsed["kpi_breakdown"]["FB_EN_ENGAGEMENT"]["operator"]

    # User says: weight saves and shares 3× (high-intent signals),
    # reactions 0.5× (low-intent), comments unchanged.
    weights = {
        components[0]["column"]: 0.5,  # Reactions: low intent
        components[1]["column"]: 1.0,  # Comments
        components[2]["column"]: 3.0,  # Shares: high intent
        components[3]["column"]: 3.0,  # Saves: high intent
    }
    weighted = [
        float(c["value"]) * weights[c["column"]] for c in components
    ]
    expected = sum(weighted) if operator == "sum" else weighted[0]
    # 0.5*4000 + 1*1500 + 3*1500 + 3*1500 = 2000 + 1500 + 4500 + 4500 = 12500
    assert expected == pytest.approx(12500.0)


def test_composition_rationale_explains_dedup_choice():
    """The rationale text should explicitly mention why link clicks are
    excluded — that's the dedup the user needs to see."""
    from core.csv_import import get_composition
    comp = get_composition("fb", "FB_EN_ENGAGEMENT")
    assert comp is not None
    text = comp.rationale.lower()
    assert "link click" in text or "double-count" in text


def test_csv_meta_export_with_total_row():
    """Meta sometimes adds a 'Results from ad sets' summary row; ensure
    that doesn't double-count."""
    csv = (
        b"Ad set name,Reach,Impressions,Amount spent\n"
        b"Set A,100000,200000,1000\n"
        b"Set B,150000,300000,1500\n"
        b'"Total results from ad sets",250000,500000,2500\n'
    )
    result = parse_platform_csv(csv, "fb")
    assert result["budget"] == pytest.approx(2500.0)
    assert result["kpis"]["FB_AW_REACH"] == pytest.approx(250000.0)


def test_resolve_then_montecarlo_uses_new_policy():
    """After re-solve with new policy, Monte Carlo should run against the
    new state — not the old one."""
    from tests.smoke_test import _run_pipeline_to_module5
    s = _run_pipeline_to_module5()
    initial_mc = run_module5_montecarlo(s, n_trials=20, seed=1)
    initial_total = sum(x.mean for x in initial_mc.per_platform)

    # Lower the budget by half, re-solve
    s.total_budget = float(s.total_budget) * 0.5
    s.module4_finalised = False
    s.module5_finalised = False
    s.module6_finalised = False
    s.current_step = 4
    run_module4(s)
    run_module5(s)

    new_mc = run_module5_montecarlo(s, n_trials=20, seed=1)
    new_total = sum(x.mean for x in new_mc.per_platform)
    # New total spend must be smaller (~half) since budget halved
    assert new_total < initial_total * 0.75, (
        f"Halved budget should halve MC total; got "
        f"{new_total:.0f} vs initial {initial_total:.0f}"
    )


def test_kpi_var_in_csv_for_unprioritised_goal_silently_dropped():
    """If a CSV carries data for a goal that Module 2 didn't prioritise on
    that platform, Module 3 should drop it — not error.  This is a normal
    case: the user exports everything Meta knows, but only AW was a
    priority on FB."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["aw", "lg"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb", "li"],
                priorities_input={
                    "fb": {"priority_1": "aw", "priority_2": None},  # FB on AW only
                    "li": {"priority_1": "lg", "priority_2": None},
                })
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "historical_days": 30, "kpis": {
            "FB_AW_REACH": 200000.0,
            "FB_LG_LEADS": 50.0,  # ← data for unprioritised goal; should be ignored
        }},
        "li": {"budget": 3000.0, "historical_days": 30, "kpis": {"LI_LG_LEADS": 80.0}},
    })
    run_module4(s)
    run_module5(s)
    # FB should serve only AW, not LG
    base = s.module5_scenario_bundle.results_by_scenario["base"]
    fb_lg = base.budget_per_platform_goal.get("fb", {}).get("lg", 0.0)
    assert fb_lg == 0.0, f"FB shouldn't get LG budget; got £{fb_lg:.0f}"


def test_module7_runs_even_with_one_scenario():
    """If only the base scenario is feasible (conservative + optimistic skipped),
    Module 7 must still produce insights — not crash on missing scenarios."""
    from modules.module5 import run_module5_lp_scenarios

    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["lg"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "lg", "priority_2": None}})
    finalise_module3_from_inputs(s, platform_inputs={
        "fb": {"budget": 3000.0, "historical_days": 30, "kpis": {"FB_LG_LEADS": 60.0}},
    })
    run_module4(s)
    # Force only a "base" scenario
    s.scenario_multipliers = {"base": 1.0}
    run_module5(s)
    run_module6(s)
    insights = run_module7(s, s.module5_scenario_bundle,
                          s.module6_scenario_result.results_by_scenario)
    assert "base" in insights.scenario_insights


def test_pipeline_with_only_rate_kpi_goal():
    """A campaign whose only goal is engagement (rate-only KPI) should
    still produce a working plan — no division by zero, no kind confusion."""
    s = WizardState()
    complete_module1_and_advance(s, raw_objectives=["en"], raw_budget=10000.0,
                                 raw_duration_days=30)
    run_module2(s, selected_platforms=["ig", "tt"],
                priorities_input={
                    "ig": {"priority_1": "en", "priority_2": None},
                    "tt": {"priority_1": "en", "priority_2": None},
                })
    finalise_module3_from_inputs(s, platform_inputs={
        "ig": {"budget": 3000.0, "historical_days": 60, "kpis": {"IG_EN_ENGRATERATE": 0.045}},
        "tt": {"budget": 3000.0, "historical_days": 60, "kpis": {"TT_EN_ENGRATERATE": 0.06}},
    })
    run_module4(s)
    run_module5(s)
    run_module6(s)
    base = s.module5_scenario_bundle.results_by_scenario["base"]
    assert base.total_budget_used > 0
    # Rate-only forecast must not multiply by budget
    base_fc = s.module6_scenario_result.results_by_scenario["base"]
    for row in base_fc.rows:
        if row.kpi_kind == "rate":
            assert row.predicted_kpi <= 1.0


def test_seasonality_doesnt_corrupt_goal_with_no_kpi_data():
    """Seasonality multiplier for a goal that has no historical data
    shouldn't blow up r_pg — multiplying zero is zero."""
    s = WizardState()
    complete_module1_and_advance(
        s, raw_objectives=["aw", "lg"], raw_budget=10000.0, raw_duration_days=30,
        raw_seasonality_index={"aw": 2.5, "lg": 0.6},
    )
    run_module2(s, selected_platforms=["fb"],
                priorities_input={"fb": {"priority_1": "aw", "priority_2": "lg"}})
    finalise_module3_from_inputs(s, platform_inputs={
        # AW data only; LG has no KPIs
        "fb": {"budget": 3000.0, "historical_days": 30, "kpis": {"FB_AW_REACH": 200000.0}},
    })
    run_module4(s)
    run_module5(s)
    base = s.module5_scenario_bundle.results_by_scenario["base"]
    # Should still solve; AW gets the budget; LG cell stays zero
    assert sum(base.budget_per_platform_goal.get("fb", {}).values()) > 0


def test_module6_rate_kpi_band_is_point_estimate():
    """Rate KPIs (engagement rate) should NOT get a ± band — they're already
    averages.  predicted_kpi_low / _high should equal predicted_kpi."""
    from tests.smoke_test import _run_pipeline_to_module5
    s = _run_pipeline_to_module5()
    run_module6(s)
    base_fc = s.module6_scenario_result.results_by_scenario["base"]
    rate_rows = [r for r in base_fc.rows if r.kpi_kind == "rate"]
    for row in rate_rows:
        assert row.predicted_kpi_low == row.predicted_kpi == row.predicted_kpi_high


def test_usd_plan_excel_and_pdf_outputs_contain_no_pound_symbol() -> None:
    """End-to-end regression: a USD plan must not leak £ into the
    rendered Excel summary, the budget allocation tables, or the PDF.
    Covers the audit's 'currency selector exists but UI still shows £
    everywhere' complaint with a check that fails loudly if any
    code path bypasses money() and hardcodes £.
    """
    import streamlit as st
    from app import (
        money,
        build_budget_allocation_df,
        build_platform_totals_df,
        create_excel_bytes,
        create_pdf_bytes,
    )
    from modules.module2 import run_module2
    from modules.module3 import finalise_module3_from_inputs
    from modules.module4 import run_module4
    from modules.module5 import run_module5
    from modules.module6 import run_module6

    s = WizardState()
    complete_module1_and_advance(
        s,
        raw_objectives=["aw", "lg"],
        raw_budget=12000.0,
        raw_currency="USD",
        raw_duration_days=30,
    )
    run_module2(
        s,
        selected_platforms=["fb", "li"],
        priorities_input={
            "fb": {"priority_1": "aw", "priority_2": None},
            "li": {"priority_1": "lg", "priority_2": None},
        },
    )
    finalise_module3_from_inputs(
        s,
        platform_inputs={
            "fb": {"budget": 4000.0, "kpis": {"FB_AW_REACH": 200000.0}},
            "li": {"budget": 3000.0, "kpis": {"LI_LG_LEADS": 80.0}},
        },
    )
    run_module4(s)
    run_module5(s)
    run_module6(s)

    # Make the session state visible to money() and friends
    st.session_state["wizard_state"] = s

    try:
        # Direct formatter check
        assert money(1234.56) == "$1,234.56", "money() should resolve to $ from session state."

        # Allocation / totals tables don't pre-format money columns themselves
        # (they're formatted at render time via money()), so this asserts the
        # consumer of those tables will see $ — verifying the upstream Df is
        # numeric (not pre-baked with a £ sign).
        budget_df = build_budget_allocation_df(s.module5_scenario_bundle.results_by_scenario["base"])
        assert "Allocated Budget" in budget_df.columns
        for v in budget_df["Allocated Budget"]:
            assert isinstance(v, float)  # raw numeric, no currency baked in

        totals_df = build_platform_totals_df(s.module5_scenario_bundle.results_by_scenario["base"])
        assert "Total Allocated Budget" in totals_df.columns
        for v in totals_df["Total Allocated Budget"]:
            assert isinstance(v, float)

        # Excel — read back via openpyxl so we only check user-visible cell
        # values, not XML namespace boilerplate.  Excel stores money columns
        # as raw floats so the user can apply their own number format in
        # Excel; the test guarantee is that no cell baked in £ via a string.
        lp = s.module5_scenario_bundle.results_by_scenario["base"]
        fc = s.module6_scenario_result.results_by_scenario["base"]
        xlsx = create_excel_bytes([("base", lp, fc)])

        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        assert "£" not in cell, (
                            f"USD plan should not contain £ in Excel cell "
                            f"({sheet_name}): {cell!r}"
                        )

        # PDF — money columns are pre-rendered via money() before being
        # placed in the table.  Extract the rendered text via pypdf rather
        # than searching the raw bytes: ReportLab uses WinAnsi single-byte
        # encoding inside content streams, so the UTF-8 byte pattern for £
        # (\xc2\xa3) never matches even when £ visually appears.  Extracted
        # text gives the symbol back in its decoded form, which is the
        # check we actually want.
        pdf = create_pdf_bytes(s, [("base", lp, fc)])

        import io
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf))
        pdf_text = "".join((p.extract_text() or "") for p in reader.pages)

        assert "$" in pdf_text, (
            "PDF export of a USD plan should contain at least one $ symbol "
            "in the rendered text."
        )
        assert "£" not in pdf_text, (
            f"PDF export of a USD plan must not contain a £ symbol in the "
            f"rendered text.  Found in text starting with: "
            f"{pdf_text[max(0, pdf_text.index('£')-40):pdf_text.index('£')+40]!r}"
            if "£" in pdf_text else
            "PDF export of a USD plan must not contain a £ symbol in the rendered text."
        )
    finally:
        # Cleanup so subsequent tests aren't affected
        if "wizard_state" in st.session_state:
            del st.session_state["wizard_state"]


def test_eur_plan_excel_output_contains_no_pound_symbol() -> None:
    """Same regression as the USD test, but for €.  We exercise both
    non-default currencies so a future regression that hardcoded $
    instead of routing through state would still fail this suite.
    """
    import streamlit as st
    from app import money, create_excel_bytes
    from modules.module2 import run_module2
    from modules.module3 import finalise_module3_from_inputs
    from modules.module4 import run_module4
    from modules.module5 import run_module5
    from modules.module6 import run_module6

    s = WizardState()
    complete_module1_and_advance(
        s,
        raw_objectives=["lg"],
        raw_budget=8000.0,
        raw_currency="EUR",
        raw_duration_days=30,
    )
    run_module2(
        s,
        selected_platforms=["li"],
        priorities_input={"li": {"priority_1": "lg", "priority_2": None}},
    )
    finalise_module3_from_inputs(
        s,
        platform_inputs={"li": {"budget": 5000.0, "kpis": {"LI_LG_LEADS": 100.0}}},
    )
    run_module4(s)
    run_module5(s)
    run_module6(s)

    st.session_state["wizard_state"] = s
    try:
        assert money(1234.56) == "€1,234.56"
        lp = s.module5_scenario_bundle.results_by_scenario["base"]
        fc = s.module6_scenario_result.results_by_scenario["base"]
        xlsx = create_excel_bytes([("base", lp, fc)])

        # Raw .xlsx bytes are a ZIP of XML files whose internals contain $
        # and similar characters in namespace boilerplate.  Read cell values
        # back via openpyxl so only user-visible content is checked.
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        assert "£" not in cell, (
                            f"EUR plan should not contain £ in Excel cell "
                            f"({sheet_name}): {cell!r}"
                        )
                        assert "$" not in cell, (
                            f"EUR plan should not contain $ in Excel cell "
                            f"({sheet_name}): {cell!r}"
                        )
    finally:
        if "wizard_state" in st.session_state:
            del st.session_state["wizard_state"]


def test_module1_error_examples_dont_anchor_on_pound_symbol() -> None:
    """Budget parse errors should reference example values in a
    currency-neutral way — a USD user shouldn't see a £ example
    when their input fails to parse.  The previous error text said
    'for example: 1200 or £1,200.50'; the new text lists all three
    accepted symbols so no currency is privileged.
    """
    from modules.module1 import _parse_budget, Module1ValidationError

    with pytest.raises(Module1ValidationError) as excinfo:
        _parse_budget("not a number")
    err = str(excinfo.value)
    # Currency-neutral: the example must list all three accepted symbols
    # rather than privileging £.  Normalise non-breaking spaces to
    # regular spaces before comparison so a formatter change that swaps
    # them wouldn't silently bypass the check.
    normalised = err.replace(" ", " ")
    assert "£, $, €" in normalised, (
        f"Expected the error to list all three accepted currency symbols, got: {err!r}"
    )
    # And the value example itself should be plain (no £ prefix on the numeric example)
    assert "£1,200.50" not in err


# ─────────────────────────────────────────────────────────────────────────────
# Unified xlsx template (one workbook → many platforms)
# ─────────────────────────────────────────────────────────────────────────────


def _platform_display_names_for_test():
    """Minimal display-names map mirroring app.py PLATFORM_NAMES, used so
    the template tests don't import app.py (which pulls in streamlit)."""
    return {
        "fb": "Facebook", "ig": "Instagram", "li": "LinkedIn",
        "yt": "YouTube", "tt": "TikTok", "pt": "Pinterest",
        "tw": "X (Twitter)", "sn": "Snapchat", "rd": "Reddit",
        "go_search":  "Google Search",
        "go_display": "Google Display",
        "go_pmax":    "Google Performance Max",
    }


def test_unified_template_one_sheet_per_selected_platform():
    """The unified xlsx template includes one sheet per supported platform
    passed in, plus an Instructions sheet at the front.  Unsupported
    platforms (none of those exist today, but defend anyway) are skipped."""
    import io
    from openpyxl import load_workbook
    from core.csv_import import generate_unified_template_xlsx

    names = _platform_display_names_for_test()
    xlsx = generate_unified_template_xlsx(
        ["fb", "li", "go_search", "go_pmax"], platform_display_names=names
    )
    assert xlsx, "Generator returned empty bytes for a valid platform list"

    wb = load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames[0] == "Instructions"
    # Subsequent sheets are in the order the platforms were passed
    assert wb.sheetnames[1:] == ["Facebook", "LinkedIn", "Google Search", "Google Performance Max"]


def test_unified_template_each_sheet_has_only_headers():
    """An unfilled sheet must contain only the header row — no example row
    that would otherwise be misread as user data when the workbook is
    parsed.  This is the invariant that lets the parser skip unfilled
    sheets unambiguously."""
    import io
    from openpyxl import load_workbook
    from core.csv_import import generate_unified_template_xlsx

    names = _platform_display_names_for_test()
    xlsx = generate_unified_template_xlsx(["fb", "li"], platform_display_names=names)
    wb = load_workbook(io.BytesIO(xlsx))

    for sheet_name in ("Facebook", "LinkedIn"):
        ws = wb[sheet_name]
        # Row 1 = headers, row 2 onwards should be empty
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 1, f"{sheet_name} has no header row"
        # All cells from row 2 onwards must be None or empty
        for row_idx, row in enumerate(rows[1:], start=2):
            assert all(c is None or str(c).strip() == "" for c in row), (
                f"{sheet_name} row {row_idx} should be empty but contains: {row}"
            )


def test_unified_template_round_trip_with_partial_fill():
    """Round-trip: generate a template, fill in only one platform's data,
    parse the workbook back.  The parser should return data for the
    filled platform and silently skip the unfilled ones."""
    import io
    from openpyxl import load_workbook
    from core.csv_import import (
        generate_unified_template_xlsx,
        parse_unified_template_xlsx,
    )

    names = _platform_display_names_for_test()
    xlsx = generate_unified_template_xlsx(
        ["fb", "li", "go_search"], platform_display_names=names
    )
    wb = load_workbook(io.BytesIO(xlsx))

    # Fill ONLY the Facebook sheet with real values
    ws = wb["Facebook"]
    header = {c.value: c.column for c in ws[1] if c.value}
    fb_data = {
        "Reach": 200000,
        "Impression": 500000,
        "Post Reactions": 1500,
        "Comments": 400,
        "Shares": 150,
        "Saves": 50,
        "Link Click": 4000,
        "On-facebook Lead": 80,
        "Amount Spent": 3000,
        "Number Of Days": 30,
    }
    for col_name, value in fb_data.items():
        if col_name in header:
            ws.cell(row=2, column=header[col_name], value=str(value))

    buf = io.BytesIO()
    wb.save(buf)
    results = parse_unified_template_xlsx(buf.getvalue(), platform_display_names=names)

    # Strip metadata keys
    parsed = {k: v for k, v in results.items() if not k.startswith("__")}

    assert list(parsed.keys()) == ["fb"], (
        f"Expected only 'fb' (the filled sheet); got {list(parsed.keys())}.  "
        "LinkedIn and Google Search were left as headers-only and should "
        "have been skipped silently."
    )

    fb = parsed["fb"]
    assert fb["budget"] == pytest.approx(3000.0)
    assert fb["kpis"]["FB_AW_REACH"] == pytest.approx(200000.0)
    assert fb["kpis"]["FB_AW_IMPRESSION"] == pytest.approx(500000.0)
    # Engagement composed from the four components: 1500 + 400 + 150 + 50 = 2100
    assert fb["kpis"]["FB_EN_ENGAGEMENT"] == pytest.approx(2100.0)
    assert fb["kpis"]["FB_WT_CLICKS"] == pytest.approx(4000.0)
    assert fb["kpis"]["FB_LG_LEADS"] == pytest.approx(80.0)


def test_unified_template_reports_unknown_sheets():
    """Sheets whose names don't match any selected platform's display
    name are listed under __unknown_sheets__ so the caller can surface
    'looks like you renamed a sheet' to the user."""
    import io
    from openpyxl import Workbook
    from core.csv_import import parse_unified_template_xlsx

    names = _platform_display_names_for_test()

    # Build a workbook by hand with a renamed sheet
    wb = Workbook()
    wb.active.title = "Instructions"
    ws = wb.create_sheet("Faceboook")  # typo
    ws.append(["Reach", "Amount Spent"])
    ws.append(["100000", "1000"])

    buf = io.BytesIO()
    wb.save(buf)
    results = parse_unified_template_xlsx(buf.getvalue(), platform_display_names=names)

    parsed = {k: v for k, v in results.items() if not k.startswith("__")}
    assert parsed == {}, (
        f"No sheet matches a known platform name; expected empty parse, got {parsed}"
    )
    assert "__unknown_sheets__" in results
    assert "Faceboook" in results["__unknown_sheets__"]["sheets"]


def test_unified_template_skips_supported_platforms_filter():
    """Passing platforms not in the CSV-supported set (e.g. a future
    addition that hasn't been wired into csv_import yet) skips them
    silently rather than crashing or producing empty sheets."""
    import io
    from openpyxl import load_workbook
    from core.csv_import import generate_unified_template_xlsx

    names = _platform_display_names_for_test()
    xlsx = generate_unified_template_xlsx(
        ["fb", "made_up_platform", "li"], platform_display_names=names
    )
    wb = load_workbook(io.BytesIO(xlsx))
    # made_up_platform silently skipped
    assert wb.sheetnames == ["Instructions", "Facebook", "LinkedIn"]


def test_unified_template_corrupt_xlsx_returns_error():
    """A truncated or non-xlsx byte stream should produce a readable
    error rather than crashing the UI."""
    from core.csv_import import parse_unified_template_xlsx
    result = parse_unified_template_xlsx(b"not a valid xlsx file")
    assert "__error__" in result
    assert "error" in result["__error__"]


def test_fb_engagement_includes_follows_component():
    """The expanded FB engagement composition adds Follows as a fifth
    summed component — Reactions + Comments + Shares + Saves + Follows.
    Locks the schema extension so a future refactor doesn't silently
    drop Follows back out."""
    from core.csv_import import parse_platform_csv
    csv = (
        b"Reach,Impression,Post Reactions,Comments,Shares,Saves,Follows,"
        b"Link Click,On-facebook Lead,Amount Spent\n"
        b"100000,200000,1000,400,200,100,50,500,20,1000\n"
    )
    result = parse_platform_csv(csv, "fb")
    # 1000 + 400 + 200 + 100 + 50 = 1750
    assert result["kpis"]["FB_EN_ENGAGEMENT"] == pytest.approx(1750.0), (
        "FB engagement should sum Reactions+Comments+Shares+Saves+Follows; "
        f"got {result['kpis']['FB_EN_ENGAGEMENT']} (expected 1750)"
    )


def test_fb_clicks_uses_first_not_sum_to_avoid_double_count():
    """FB now surfaces Link Click, Landing Page View, and Page View as
    three separate template columns under FB_WT_CLICKS.  But Landing
    Page Views are a subset of Link Clicks (clicks that successfully
    loaded), so summing would double-count.  operator='first' picks the
    most-canonical Link Click value when more than one is filled."""
    from core.csv_import import parse_platform_csv
    csv = (
        b"Reach,Link Click,Landing Page View,Page View,Amount Spent\n"
        b"100000,4000,3500,2000,1000\n"
    )
    result = parse_platform_csv(csv, "fb")
    # Link Click wins (first non-zero), NOT 4000+3500+2000=9500
    assert result["kpis"]["FB_WT_CLICKS"] == pytest.approx(4000.0), (
        f"Expected Link Click (4000) to win over Landing/Page View alternates; "
        f"got {result['kpis']['FB_WT_CLICKS']}"
    )


def test_fb_conversions_first_not_sum_to_avoid_triple_count():
    """FB_LG_LEADS now accepts Leads, Purchases, or Conversions as
    alternates.  Meta's 'Conversions' is usually a superset that
    already includes Leads and Purchases — summing would triple-count.
    Verify the parser picks one canonical value."""
    from core.csv_import import parse_platform_csv
    csv = (
        b"Reach,On-facebook Lead,Purchases,Conversions,Amount Spent\n"
        b"100000,80,40,150,1000\n"
    )
    result = parse_platform_csv(csv, "fb")
    # Leads (80) wins; NOT 80+40+150=270
    assert result["kpis"]["FB_LG_LEADS"] == pytest.approx(80.0), (
        f"Expected Lead (80) to win over Purchases/Conversions; "
        f"got {result['kpis']['FB_LG_LEADS']} (would be 270 if summed)"
    )


def test_new_platforms_have_csv_patterns_pt_tw_sn_rd():
    """The four platforms that previously had KPI_CONFIG entries but no
    CSV-import patterns (Pinterest, X, Snapchat, Reddit) now have full
    parse support."""
    from core.csv_import import SUPPORTED_PLATFORMS
    for p in ("pt", "tw", "sn", "rd"):
        assert p in SUPPORTED_PLATFORMS, f"{p!r} missing from CSV-import catalogue"


def test_pinterest_parses_full_native_schema():
    """Smoke-test Pinterest's new CSV patterns: Impression/Video View
    for awareness, Saves for engagement, Outbound/Pin Click for traffic,
    Leads/Checkouts for conversion."""
    from core.csv_import import parse_platform_csv
    csv = (
        b"Impression,Video View,Saves,Outbound Click,Pin Click,Leads,Checkouts,Cost\n"
        b"80000,30000,600,700,400,25,10,800\n"
    )
    result = parse_platform_csv(csv, "pt")
    assert result["budget"] == pytest.approx(800.0)
    assert result["kpis"]["PT_AW_IMPRESSION"] == pytest.approx(80000.0)
    assert result["kpis"]["PT_EN_SAVES"] == pytest.approx(600.0)
    assert result["kpis"]["PT_WT_CLICKS"] == pytest.approx(700.0)  # Outbound preferred
    assert result["kpis"]["PT_LG_LEADS"] == pytest.approx(25.0)


def test_x_rate_kpi_normalised_from_percent_form():
    """X engagement rate must normalise '3.5%' → 0.035, same as the
    other rate-canonical platforms.  TW_EN_ENGRATERATE is in _RATE_KPIS."""
    from core.csv_import import parse_platform_csv
    csv = b"Impression,Engagement Rate,Link Click,Leads,Cost\n100000,3.5%,1200,50,1000\n"
    result = parse_platform_csv(csv, "tw")
    assert result["kpis"]["TW_EN_ENGRATERATE"] == pytest.approx(0.035), (
        f"X engagement rate should be normalised to [0,1]; got "
        f"{result['kpis']['TW_EN_ENGRATERATE']}"
    )


def test_template_informational_extras_appear_but_dont_pollute_canonicals():
    """Rate-canonical engagement platforms (IG, TT, etc.) surface raw
    count columns (Likes, Comments, Shares, Saves, Follows) in the
    template for the user's records, but those columns must NOT feed
    the canonical engagement rate.  If the user fills Likes but not
    Engagement Rate, the canonical stays missing rather than producing
    a corrupt 'rate' from a count."""
    import io
    from openpyxl import load_workbook
    from core.csv_import import (
        generate_unified_template_xlsx,
        parse_unified_template_xlsx,
    )

    names = _platform_display_names_for_test()
    xlsx = generate_unified_template_xlsx(["ig"], platform_display_names=names)
    wb = load_workbook(io.BytesIO(xlsx))

    # Confirm the informational columns appear in the IG sheet header
    ig_columns = [c.value for c in wb["Instagram"][1]]
    for extra in ("Likes", "Comments", "Shares", "Saves", "Follows"):
        assert extra in ig_columns, (
            f"Informational column {extra!r} missing from Instagram template"
        )

    # Fill only the informational counts, NOT the rate or any canonical
    ws = wb["Instagram"]
    hdr = {c.value: c.column for c in ws[1]}
    counts = {
        "Reach": 100000, "Amount Spent": 1000, "Number Of Days": 30,
        "Likes": 5000, "Comments": 200, "Shares": 100, "Saves": 80, "Follows": 30,
    }
    for col, val in counts.items():
        if col in hdr:
            ws.cell(row=2, column=hdr[col], value=str(val))

    buf = io.BytesIO()
    wb.save(buf)
    results = parse_unified_template_xlsx(buf.getvalue(), platform_display_names=names)
    parsed = {k: v for k, v in results.items() if not k.startswith("__")}

    assert "ig" in parsed
    # The rate KPI should be missing (user didn't fill it) — Likes shouldn't
    # have been mis-parsed AS the engagement rate.
    assert "IG_EN_ENGRATERATE" not in parsed["ig"]["kpis"], (
        f"Informational Likes column polluted IG_EN_ENGRATERATE; got "
        f"{parsed['ig']['kpis'].get('IG_EN_ENGRATERATE')}"
    )
    # But the canonicals that DID match should be present
    assert parsed["ig"]["kpis"]["IG_AW_REACH"] == pytest.approx(100000.0)
