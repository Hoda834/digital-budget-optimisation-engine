"""CSV import + canonical-KPI composition layer.

Marketers report different metrics on each platform — Facebook's
"engagement" is a different beast from LinkedIn's, and that's a different
beast from TikTok's.  The LP needs *one* number per (platform, goal) cell,
so something has to compose the raw platform metrics into the canonical
category the LP optimises against.

This module makes that composition explicit instead of inheriting
whichever number the platform's CSV happens to bundle.  Each canonical
KPI declares:

  - one or more *components* (alternative column names that all measure
    the same atomic signal — e.g. "Reactions" or "Post reactions");
  - an *operator* combining the components (sum / first / max / mean);
  - a *rationale* explaining the choice, including any deduplication
    against other canonical categories.

For example, FB_EN_ENGAGEMENT is the sum of reactions + comments +
shares + saves — explicitly excluding link clicks because those are in
FB_WT_CLICKS and would otherwise be double-counted.

The composition is auditable: every parsed CSV returns a kpi_breakdown
showing which raw columns went into each canonical KPI, with the
operator that combined them, so the user can see exactly what produced
the LP's input.  Users can override the value directly in the form, or
re-weight components via the optional "Customise composition" panel.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union


# ─────────────────────────────────────────────────────────────────────────────
# Composition rules
# ─────────────────────────────────────────────────────────────────────────────


@dataclass(frozen=True)
class KPIComposition:
    """How a canonical KPI is built from raw CSV columns.

    components: list of component groups.  Each group is a tuple of
        alternate column-name needles — the first column matching any
        needle in the group provides the component's value.  Different
        groups produce different components that get combined.

    operator: how to combine component values across groups.
        - 'first': take the first non-zero component value (used when
                   the needles in a single group are alternatives, e.g.
                   "leads" OR "lead form submissions").
        - 'sum':   add component values (used for genuine composites
                   like engagement = reactions + comments + shares).
        - 'max':   take the largest component value (e.g. reach vs
                   impressions when both are reported).
        - 'mean':  average component values (used for rates).

    rationale: short human-readable text the UI surfaces to explain
        why this composition was chosen and which atomic signals it
        deliberately excludes to avoid double-counting.

    fallback: a second composition to try if the primary one matches no
        columns.  Used when the platform offers both a clean break-out
        ("Reactions", "Comments", ...) and a bundled rollup ("Post
        engagement").  The rollup is usually less safe — it can include
        signals that already live in another canonical category — so the
        fallback rationale should warn the user about that.
    """
    components: Tuple[Tuple[str, ...], ...]
    operator: str = "first"
    rationale: str = ""
    fallback: Optional["KPIComposition"] = None


# Pseudo-KPI markers for non-KPI columns we still want from the CSV.
_BUDGET = "_budget"
_DAYS = "_days"


_CSV_PATTERNS: Dict[str, Dict[str, KPIComposition]] = {
    # ── Meta / Facebook ────────────────────────────────────────────────────
    "fb": {
        "FB_AW_REACH": KPIComposition(
            components=(("reach",),), operator="first",
            rationale="Unique people who saw the ad. Distinct from impressions.",
        ),
        "FB_AW_IMPRESSION": KPIComposition(
            # Impressions are total served; Video Views are a count of
            # video plays — both are awareness signals that overlap, so
            # we surface both as template columns but take FIRST to
            # avoid double-counting if the user reports both.
            components=(("impression",), ("video view", "video views")),
            operator="first",
            rationale="Total impressions served (includes repeats). Video Views accepted as a fallback.",
        ),
        "FB_EN_ENGAGEMENT": KPIComposition(
            components=(
                ("post reactions", "reactions"),
                ("comments", "post comments"),
                ("shares", "post shares"),
                ("saves",),
                ("follows", "page follows", "new follows"),
            ),
            operator="sum",
            rationale=(
                "Reactions + comments + shares + saves + follows.  Link clicks "
                "are excluded — they live under Website Traffic, so including "
                "them here would double-count."
            ),
            fallback=KPIComposition(
                components=(("post engagement", "engagement"),),
                operator="first",
                rationale=(
                    "Using Meta's bundled 'Post engagement' column because "
                    "the broken-out components weren't in your export.  "
                    "Be aware: this number includes link clicks, which are "
                    "ALSO in the Traffic category — the LP may see a small "
                    "amount of double-counting for FB.  To fix, export "
                    "reactions / comments / shares / saves / follows as "
                    "separate columns and re-upload."
                ),
            ),
        ),
        "FB_WT_CLICKS": KPIComposition(
            # Link Clicks is the canonical destination signal; Landing
            # Page Views and Page Views are alternates surfaced as
            # separate template columns.  operator='first' prevents
            # double-counting when more than one is filled (Landing Page
            # Views is a subset of Link Clicks; Page Views is on-platform).
            components=(
                ("link click", "outbound click", "click (all)"),
                ("landing page view", "landing page views"),
                ("page view", "page views"),
            ),
            operator="first",
            rationale=(
                "Clicks to your destination URL.  Link Click is preferred; "
                "Landing Page Views or Page Views accepted as fallbacks."
            ),
        ),
        "FB_LG_LEADS": KPIComposition(
            components=(
                ("on-facebook lead", "leads", "lead"),
                ("conversions", "conv."),
            ),
            operator="first",
            rationale=(
                "On-platform lead form submissions preferred; Conversions "
                "accepted as a fallback (Meta's 'Conversions' is often a "
                "superset that already includes Leads, so summing would "
                "double-count).  Purchases is now its own canonical "
                "(FB_LG_PURCHASES) so a purchase-goal campaign isn't "
                "lumped with lead-gen volume."
            ),
        ),
        "FB_LG_PURCHASES": KPIComposition(
            components=(("purchases", "purchase"),), operator="first",
            rationale=(
                "Purchase events tracked separately so the optimiser can "
                "reward purchases without conflating them with Leads."
            ),
        ),
        _BUDGET: KPIComposition(
            components=(("amount spent", "spend", "cost"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── Instagram ──────────────────────────────────────────────────────────
    "ig": {
        "IG_AW_REACH": KPIComposition(
            components=(("reach",), ("impression", "impressions"), ("reel view", "reel views")),
            operator="first",
            rationale="Unique people who saw the ad.  Impressions / Reel Views accepted as fallbacks.",
        ),
        "IG_EN_ENGAGEMENT": KPIComposition(
            components=(
                ("likes", "like"),
                ("comments", "comment"),
                ("shares", "share"),
                ("saves", "save"),
                ("follows", "follow"),
            ),
            operator="sum",
            rationale="Likes + Comments + Shares + Saves + Follows. Engagement is a count, "
                      "so the unit matches Reach, Clicks and Leads on this platform.",
        ),
        "IG_WT_CLICKS": KPIComposition(
            components=(
                ("website click", "website clicks", "link click", "outbound click"),
                ("profile visit", "profile visits"),
            ),
            operator="first",
            rationale="Clicks to your destination URL.  Profile Visits accepted as a fallback.",
        ),
        "IG_LG_LEADS": KPIComposition(
            components=(("leads", "lead"),),
            operator="first",
            rationale="Lead form submissions.",
        ),
        "IG_LG_PURCHASES": KPIComposition(
            components=(("purchases", "purchase"),),
            operator="first",
            rationale="Purchase events tracked via the IG conversion API.",
        ),
        _BUDGET: KPIComposition(
            components=(("amount spent", "spend", "cost"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── LinkedIn ───────────────────────────────────────────────────────────
    "li": {
        "LI_AW_REACH": KPIComposition(
            components=(("impression",), ("reach",), ("video view", "video views")),
            operator="first",
            rationale="LinkedIn reports Impressions; Reach / Video Views accepted as fallbacks.",
        ),
        "LI_EN_ENGAGEMENT": KPIComposition(
            components=(
                ("reactions", "reaction"),
                ("comments", "comment"),
                ("shares", "share"),
                ("followers", "follower"),
            ),
            operator="sum",
            rationale="Reactions + Comments + Shares + Followers. Count units, "
                      "matching the other LinkedIn canonicals.",
        ),
        "LI_WT_CLICKS": KPIComposition(
            components=(("click",), ("website visit", "website visits")),
            operator="first",
            rationale="Clicks on the ad.  Website Visits accepted as a fallback.",
        ),
        "LI_LG_LEADS": KPIComposition(
            components=(("leads", "lead"), ("conversions", "conv.")),
            operator="first",
            rationale="Lead Gen Form completions preferred; Conversions accepted as a fallback.",
        ),
        _BUDGET: KPIComposition(
            components=(("total spent", "amount spent", "spend", "cost"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── Google Search ──────────────────────────────────────────────────────
    # No engagement KPI — see KPI_CONFIG comment.  All Google canonicals
    # are counts (Impressions, Clicks, Conversions, Purchases) to preserve
    # the platform-uniform-units invariant.
    "go_search": {
        "GO_SEARCH_AW_IMPRESSION": KPIComposition(
            components=(("impr.", "impressions", "impression"),), operator="first",
            rationale="Search-only impressions (export filtered to Search campaigns).",
        ),
        "GO_SEARCH_WT_CLICKS": KPIComposition(
            components=(("clicks", "click"),), operator="first",
            rationale="Clicks on Search ads.",
        ),
        "GO_SEARCH_LG_CONVERSIONS": KPIComposition(
            components=(
                ("conversions", "conv."),
                ("leads", "lead"),
                ("calls", "phone calls"),
            ),
            operator="first",
            rationale="Conversions preferred (whatever event you tagged); "
                      "Leads / Calls accepted as fallbacks.  Purchases is now "
                      "its own canonical (GO_SEARCH_LG_PURCHASES).",
        ),
        "GO_SEARCH_LG_PURCHASES": KPIComposition(
            components=(("purchases", "purchase"),), operator="first",
            rationale="Purchase conversions tracked separately so a purchase-goal "
                      "campaign isn't lumped with lead-gen Conversions.",
        ),
        _BUDGET: KPIComposition(
            components=(("cost", "spend"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── Google Display Network ─────────────────────────────────────────────
    "go_display": {
        "GO_DISPLAY_AW_IMPRESSION": KPIComposition(
            components=(("impr.", "impressions", "impression"), ("view", "views")),
            operator="first",
            rationale="Display Network impressions; Views accepted as a fallback.",
        ),
        "GO_DISPLAY_WT_CLICKS": KPIComposition(
            components=(("clicks", "click"),), operator="first",
            rationale="Clicks on Display ads (including responsive display).",
        ),
        "GO_DISPLAY_LG_CONVERSIONS": KPIComposition(
            components=(("conversions", "conv."),), operator="first",
            rationale="Conversions from Display campaigns.",
        ),
        "GO_DISPLAY_LG_PURCHASES": KPIComposition(
            components=(("purchases", "purchase"),), operator="first",
            rationale="Purchase conversions tracked separately.",
        ),
        _BUDGET: KPIComposition(
            components=(("cost", "spend"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── Google Performance Max ─────────────────────────────────────────────
    "go_pmax": {
        "GO_PMAX_AW_IMPRESSION": KPIComposition(
            components=(("impr.", "impressions", "impression"), ("view", "views")),
            operator="first",
            rationale="PMax impressions blended across surfaces; Views accepted as a fallback.",
        ),
        "GO_PMAX_WT_CLICKS": KPIComposition(
            components=(("clicks", "click"),), operator="first",
            rationale="Total PMax clicks across all surfaces.",
        ),
        "GO_PMAX_LG_CONVERSIONS": KPIComposition(
            components=(
                ("conversions", "conv."),
                ("leads", "lead"),
                ("store visit", "store visits"),
            ),
            operator="first",
            rationale="Conversions preferred (Smart Bidding optimises for it); "
                      "Leads / Store Visits accepted as fallbacks.  Purchases "
                      "is now its own canonical (GO_PMAX_LG_PURCHASES).",
        ),
        "GO_PMAX_LG_PURCHASES": KPIComposition(
            components=(("purchases", "purchase"),), operator="first",
            rationale="Purchase conversions tracked separately.",
        ),
        _BUDGET: KPIComposition(
            components=(("cost", "spend"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── TikTok ─────────────────────────────────────────────────────────────
    "tt": {
        "TT_AW_VIEWS": KPIComposition(
            components=(("video views", "views"), ("reach",)),
            operator="first",
            rationale="Total video views (2 seconds+); Reach accepted as a fallback.",
        ),
        "TT_EN_ENGAGEMENT": KPIComposition(
            components=(
                ("likes", "like"),
                ("comments", "comment"),
                ("shares", "share"),
                ("saves", "save"),
                ("followers", "follower"),
            ),
            operator="sum",
            rationale="Likes + Comments + Shares + Saves + Followers. Count units.",
        ),
        "TT_WT_CLICKS": KPIComposition(
            components=(("destination click", "clicks"), ("profile view", "profile views")),
            operator="first",
            rationale="Destination clicks preferred; Profile Views accepted as a fallback.",
        ),
        "TT_LG_LEADS": KPIComposition(
            components=(("leads", "lead"),),
            operator="first",
            rationale="Lead form submissions.",
        ),
        "TT_LG_PURCHASES": KPIComposition(
            components=(("purchases", "purchase"),),
            operator="first",
            rationale="Purchase events tracked separately.",
        ),
        _BUDGET: KPIComposition(
            components=(("cost", "spend"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── YouTube ────────────────────────────────────────────────────────────
    "yt": {
        "YT_AW_VIEWS": KPIComposition(
            components=(("views",), ("impression", "impressions"), ("unique viewer", "unique viewers")),
            operator="first",
            rationale="Total ad views preferred; Impressions / Unique Viewers accepted as fallbacks.",
        ),
        "YT_EN_ENGAGEMENT": KPIComposition(
            components=(
                ("likes", "like"),
                ("comments", "comment"),
                ("shares", "share"),
                ("subscribers", "subscriber"),
            ),
            operator="sum",
            rationale="Likes + Comments + Shares + Subscribers. Count units, "
                      "replacing the legacy View-Rate canonical.",
        ),
        "YT_WT_CLICKS": KPIComposition(
            components=(
                ("clicks", "click"),
                ("card click", "card clicks"),
                ("end screen click", "end screen clicks"),
            ),
            operator="first",
            rationale="Total clicks preferred; Card / End Screen clicks accepted as fallbacks.",
        ),
        "YT_LG_LEADS": KPIComposition(
            components=(("conversions", "leads"),),
            operator="first",
            rationale="Tagged conversions (called Conversions on YouTube).",
        ),
        "YT_LG_PURCHASES": KPIComposition(
            components=(("purchases", "purchase"),),
            operator="first",
            rationale="Purchase events tracked separately.",
        ),
        _BUDGET: KPIComposition(
            components=(("cost", "spend"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── Pinterest ──────────────────────────────────────────────────────────
    "pt": {
        "PT_AW_IMPRESSION": KPIComposition(
            components=(("impression",), ("video view", "video views")),
            operator="first",
            rationale="Total impressions; Video Views accepted as a fallback.",
        ),
        "PT_EN_SAVES": KPIComposition(
            components=(("saves", "save"),), operator="first",
            rationale="Pin saves — the canonical Pinterest engagement signal.  "
                      "Closeups / Followers stay informational because Saves is "
                      "the labelled canonical (per KPI_CONFIG).",
        ),
        "PT_WT_CLICKS": KPIComposition(
            components=(("outbound click", "outbound clicks"), ("pin click", "pin clicks")),
            operator="first",
            rationale="Outbound clicks preferred (off-platform traffic); Pin Clicks accepted as a fallback.",
        ),
        "PT_LG_LEADS": KPIComposition(
            components=(("leads", "lead"),),
            operator="first",
            rationale="Leads.",
        ),
        "PT_LG_PURCHASES": KPIComposition(
            components=(("checkouts", "checkout"),), operator="first",
            rationale="Checkouts — Pinterest's purchase-equivalent event.",
        ),
        _BUDGET: KPIComposition(
            components=(("cost", "spend"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── X (Twitter) ────────────────────────────────────────────────────────
    "tw": {
        "TW_AW_IMPRESSION": KPIComposition(
            components=(("impression",), ("video view", "video views")),
            operator="first",
            rationale="Total impressions; Video Views accepted as a fallback.",
        ),
        "TW_EN_ENGAGEMENT": KPIComposition(
            components=(
                ("likes", "like"),
                ("replies", "reply"),
                ("reposts", "repost"),
                ("bookmarks", "bookmark"),
                ("followers", "follower"),
            ),
            operator="sum",
            rationale="Likes + Replies + Reposts + Bookmarks + Followers. Count units.",
        ),
        "TW_WT_CLICKS": KPIComposition(
            components=(("link click", "link clicks"), ("profile visit", "profile visits")),
            operator="first",
            rationale="Link clicks preferred (off-platform traffic); Profile Visits accepted as a fallback.",
        ),
        "TW_LG_LEADS": KPIComposition(
            components=(("leads", "lead"),),
            operator="first",
            rationale="Lead form submissions.",
        ),
        _BUDGET: KPIComposition(
            components=(("cost", "spend", "amount spent"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── Snapchat ───────────────────────────────────────────────────────────
    "sn": {
        "SN_AW_REACH": KPIComposition(
            components=(("reach",), ("impression", "impressions")),
            operator="first",
            rationale="Unique reach preferred; Impressions accepted as a fallback.",
        ),
        "SN_EN_ENGAGEMENT": KPIComposition(
            components=(
                ("story opens", "story open"),
                ("shares", "share"),
                ("subscribers", "subscriber"),
            ),
            operator="sum",
            rationale="Story Opens + Shares + Subscribers. Count units.",
        ),
        "SN_WT_CLICKS": KPIComposition(
            components=(("swipe-up", "swipe-ups", "swipe ups", "swipeups"),),
            operator="first",
            rationale="Swipe-ups (Snapchat's destination-click metric).",
        ),
        "SN_LG_LEADS": KPIComposition(
            components=(("leads", "lead"),),
            operator="first",
            rationale="Leads.",
        ),
        "SN_LG_PURCHASES": KPIComposition(
            components=(("purchases", "purchase"),),
            operator="first",
            rationale="Purchase events tracked separately.",
        ),
        _BUDGET: KPIComposition(
            components=(("cost", "spend", "amount spent"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
    # ── Reddit ─────────────────────────────────────────────────────────────
    "rd": {
        "RD_AW_IMPRESSION": KPIComposition(
            components=(("impression",), ("video view", "video views")),
            operator="first",
            rationale="Total impressions; Video Views accepted as a fallback.",
        ),
        "RD_EN_ENGAGEMENT": KPIComposition(
            components=(
                ("upvotes", "upvote"),
                ("comments", "comment"),
                ("shares", "share"),
                ("followers", "follower"),
            ),
            operator="sum",
            rationale="Upvotes + Comments + Shares + Followers. Count units.",
        ),
        "RD_WT_CLICKS": KPIComposition(
            components=(("clicks", "click"),),
            operator="first",
            rationale="Clicks on the ad.",
        ),
        "RD_LG_LEADS": KPIComposition(
            components=(("leads", "lead"), ("conversions", "conv.")),
            operator="first",
            rationale="Leads preferred; Conversions accepted as a fallback.",
        ),
        _BUDGET: KPIComposition(
            components=(("cost", "spend", "amount spent"),), operator="first",
        ),
        _DAYS: KPIComposition(
            components=(("number of days", "days"),), operator="first",
        ),
    },
}


# Informational template columns that aren't mapped to any canonical KPI.
# Currently empty — Pinterest's Closeups/Followers stayed informational
# in a previous iteration but are now omitted entirely since the
# PT_EN_SAVES canonical is specifically Saves and surfacing the others
# implied they fed into it (which was misleading).
#
# This map is retained so a future rate-canonical platform can register
# template-only columns without re-introducing the dead-data trap.
_TEMPLATE_EXTRA_COLUMNS: Dict[str, Tuple[Tuple[str, str], ...]] = {
    "pt": (
        ("Closeups", "3000"), ("Followers", "100"),
    ),
}


# Engagement post-processing: if a user uploads a legacy export with an
# 'Engagement Rate' column but no individual count breakouts, derive
# engagement count = rate × awareness count.  Maps each social platform
# to (set of rate-column needles, awareness canonical var to multiply
# by, target engagement canonical var).  The post-processing only fires
# when the engagement canonical came out missing from the primary
# component aggregation.
_LEGACY_RATE_FALLBACK: Dict[str, Tuple[Tuple[str, ...], str, str]] = {
    "ig": (("engagement rate", "er"),                "IG_AW_REACH",       "IG_EN_ENGAGEMENT"),
    "li": (("engagement rate", "average ctr"),       "LI_AW_REACH",       "LI_EN_ENGAGEMENT"),
    "yt": (("view rate", "engagement rate"),         "YT_AW_VIEWS",       "YT_EN_ENGAGEMENT"),
    "tt": (("engagement rate", "er"),                "TT_AW_VIEWS",       "TT_EN_ENGAGEMENT"),
    "tw": (("engagement rate", "er"),                "TW_AW_IMPRESSION",  "TW_EN_ENGAGEMENT"),
    "sn": (("engagement rate", "er"),                "SN_AW_REACH",       "SN_EN_ENGAGEMENT"),
    "rd": (("engagement rate", "er"),                "RD_AW_IMPRESSION",  "RD_EN_ENGAGEMENT"),
}


# Set of canonical KPI vars whose values are stored as a rate (decimal
# in [0,1]).  This set is now empty for the social platforms — all
# engagement KPIs are counts so the same units apply within a platform.
# Retained as an empty set for backward compatibility with code paths
# that branch on `var in _RATE_KPIS`; if a future platform reintroduces
# a rate canonical, register it here.
_RATE_KPIS: set = set()


SUPPORTED_PLATFORMS: Tuple[str, ...] = tuple(_CSV_PATTERNS.keys())


def get_composition(platform: str, var: str) -> Optional[KPIComposition]:
    """Return the KPIComposition for a (platform, var) pair, or None if
    the platform isn't in the CSV-import set or the var isn't recognised."""
    return _CSV_PATTERNS.get(platform.lower(), {}).get(var)


def _template_example_value(var: str) -> str:
    """Sample value for the template's example row.  Sizes are
    order-of-magnitude plausible for a £3k monthly campaign so the
    template doesn't suggest unrealistic ratios."""
    if var == _BUDGET:
        return "3000"
    if var == _DAYS:
        return "30"
    if var in _RATE_KPIS:
        return "2.5%"
    lower = var.lower()
    if "reach" in lower or "impression" in lower or "view" in lower:
        return "500000"
    if "click" in lower:
        return "5000"
    if "lead" in lower or "conversion" in lower:
        return "100"
    if "reaction" in lower or "comment" in lower or "share" in lower or "save" in lower:
        return "2000"
    if "engagement" in lower:
        return "8000"
    return "1000"


def _column_name_for_needle(needle: str) -> str:
    """Convert a normalised needle ('amount spent') into a presentable
    column name ('Amount Spent') for the template."""
    # Title-case but keep abbreviations and common headers natural
    parts = needle.split()
    out = []
    for p in parts:
        if p.lower() in ("ctr", "er", "cpc", "cpa", "cpm"):
            out.append(p.upper())
        else:
            out.append(p.capitalize())
    return " ".join(out)


def generate_csv_template(platform: str) -> bytes:
    """Build a downloadable CSV template for one platform.

    The header row contains every column the parser looks for (one per
    component group, not just one per canonical KPI — so engagement
    breaks out into Reactions / Comments / Shares / Saves rather than
    a single 'Engagement' bucket).  One example row follows with
    order-of-magnitude plausible values for a typical month.

    Returns empty bytes if the platform isn't in the CSV-import set.
    """
    columns, examples = _template_columns_and_examples(platform)
    if not columns:
        return b""
    csv_text = ",".join(columns) + "\n" + ",".join(examples) + "\n"
    return csv_text.encode("utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# Unified xlsx template (one workbook covering every selected platform)
# ─────────────────────────────────────────────────────────────────────────────


# Excel sheet names have a 31-character limit and forbid: \ / ? * [ ] :
# Trim and sanitise display names against that before writing.
_INVALID_SHEET_CHARS = set(r"\/?*[]:")


def _sanitise_sheet_name(name: str) -> str:
    cleaned = "".join("_" if c in _INVALID_SHEET_CHARS else c for c in str(name))
    return cleaned[:31].strip() or "Sheet"


def _template_columns_and_examples(platform: str) -> Tuple[List[str], List[str]]:
    """Return (columns, example_values) for a platform's template.

    Shared between the per-platform CSV template and the unified-workbook
    template so both surface the same column shape — every component
    group as its own column, deduplicated.
    """
    plat = (platform or "").strip().lower()
    if plat not in _CSV_PATTERNS:
        return [], []
    patterns = _CSV_PATTERNS[plat]
    columns: List[str] = []
    examples: List[str] = []
    seen: set = set()
    for var, comp in patterns.items():
        for needle_group in comp.components:
            if not needle_group:
                continue
            col_name = _column_name_for_needle(needle_group[0])
            if col_name in seen:
                continue
            seen.add(col_name)
            columns.append(col_name)
            examples.append(_template_example_value(var))
    # Informational extras — raw count columns surfaced for the user's
    # records on platforms whose engagement canonical is a rate.  The
    # parser ignores these silently; they don't contribute to any KPI.
    for col_name, example in _TEMPLATE_EXTRA_COLUMNS.get(plat, ()):
        if col_name in seen:
            continue
        seen.add(col_name)
        columns.append(col_name)
        examples.append(example)
    return columns, examples


def generate_unified_template_xlsx(
    platforms: Sequence[str],
    platform_display_names: Optional[Dict[str, str]] = None,
) -> bytes:
    """Build one Excel workbook covering every supplied platform.

    Each platform gets its own sheet named after its display label (or
    platform code if no display name is provided).  Sheet 1 is an
    instructions/index sheet listing each platform sheet and what the
    user should fill in.

    Platforms not in the CSV-import catalogue are skipped silently so a
    caller can pass the full active_platforms list without filtering.

    Returns empty bytes if openpyxl is unavailable or no platforms map
    to a supported template.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return b""

    display = platform_display_names or {}
    eligible = [p for p in (platforms or []) if (p or "").strip().lower() in _CSV_PATTERNS]
    if not eligible:
        return b""

    wb = Workbook()
    # First sheet is the active default; rename it to the instructions page.
    instructions = wb.active
    instructions.title = "Instructions"

    instructions["A1"] = "Unified historical-data template"
    instructions["A1"].font = Font(bold=True, size=14)
    instructions["A3"] = (
        "Fill in one sheet per platform you ran during the historical window. "
        "Leave a sheet completely empty (just the headers row) if you didn't "
        "run that platform — the parser silently skips empty sheets."
    )
    instructions["A4"] = "Each platform sheet has only headers — add ONE row of your own data underneath."
    instructions["A5"] = "  • 'Amount Spent' (or 'Cost') — total spend in the window you're reporting on."
    instructions["A6"] = "  • 'Number of Days' — how many days of history you're reporting on."
    instructions["A7"] = (
        "  • One column per metric the platform's native export provides "
        "(Reach, Impressions, Engagement, etc.).  Fill the columns you have; "
        "leave the rest blank."
    )
    instructions["A8"] = "Then upload this workbook back into Module 3."
    instructions["A10"] = "Realistic ranges for a £3k monthly campaign (illustrative, not required):"
    instructions["A10"].font = Font(bold=True)
    instructions["A11"] = "  • Reach / Impressions:        100,000 – 1,000,000"
    instructions["A12"] = "  • Clicks:                     1,000 – 10,000"
    instructions["A13"] = "  • Leads / Conversions:        20 – 200"
    instructions["A14"] = "  • Purchases:                  10 – 100"
    instructions["A15"] = "  • Engagement count components: 100 – 5,000 each (likes, comments, shares, etc.)"
    instructions["A17"] = "Sheets in this workbook:"
    instructions["A17"].font = Font(bold=True)

    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)

    used_sheet_names: Dict[str, str] = {}  # sheet_name → platform_code
    for i, p in enumerate(eligible, start=1):
        code = p.strip().lower()
        sheet_label = display.get(code, code)
        sheet_name = _sanitise_sheet_name(sheet_label)
        # Disambiguate if two platforms share a sanitised name (shouldn't
        # happen with the current catalogue but defend anyway)
        base = sheet_name
        suffix = 2
        while sheet_name in used_sheet_names:
            sheet_name = _sanitise_sheet_name(f"{base} {suffix}")
            suffix += 1
        used_sheet_names[sheet_name] = code

        ws = wb.create_sheet(title=sheet_name)
        # Header row only — no example row baked into the data area, so an
        # untouched sheet is unambiguously empty and the parser skips it.
        columns, _examples = _template_columns_and_examples(code)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = max(
                14, min(40, len(col_name) + 2)
            )

        instructions.cell(row=17 + i, column=1, value=f"  • {sheet_name}: {len(columns)} columns")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_unified_template_xlsx(
    xlsx_bytes: bytes,
    platform_display_names: Optional[Dict[str, str]] = None,
) -> Dict[str, Dict[str, Any]]:
    """Parse a filled-in unified template workbook.

    Returns a {platform_code: parse_result} mapping where each parse_result
    has the same shape as parse_platform_csv() — i.e. ``budget``, ``kpis``,
    ``kpi_breakdown``, ``matched_columns``, ``missing_kpis``, plus an
    optional ``error`` key.

    Sheets that don't correspond to a known platform are reported under
    the ``__unknown_sheets__`` key in the returned dict so the caller can
    surface "you renamed a sheet" warnings.  Sheets with no data row (or
    only the example row left untouched) are skipped silently.

    Mapping rule: sheet name is matched against the sanitised display
    name of each platform.  If the user renames a sheet, that sheet is
    skipped — they have to either rename it back or upload the original
    per-platform CSV through the per-platform upload widget.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"__error__": {"error": "openpyxl is required to parse the unified xlsx template."}}

    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as exc:  # invalid xlsx
        return {"__error__": {"error": f"Could not open xlsx: {exc}"}}

    display = platform_display_names or {}
    # Build the inverse: sanitised-sheet-name → platform_code
    sheet_to_code: Dict[str, str] = {}
    for code in _CSV_PATTERNS:
        label = display.get(code, code)
        sheet_to_code[_sanitise_sheet_name(label)] = code

    results: Dict[str, Dict[str, Any]] = {}
    unknown_sheets: List[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Instructions":
            continue
        code = sheet_to_code.get(sheet_name)
        if code is None:
            unknown_sheets.append(sheet_name)
            continue

        ws = wb[sheet_name]
        # Read the sheet as CSV-like rows; openpyxl gives None for empty cells.
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            rows.append(["" if c is None else str(c) for c in row])

        if len(rows) < 2:
            # Header row only — user didn't fill anything in
            continue

        # Convert to a CSV byte stream so we can reuse parse_platform_csv unchanged
        buf = io.StringIO()
        writer = csv.writer(buf)
        for r in rows:
            writer.writerow(r)
        csv_bytes = buf.getvalue().encode("utf-8")

        results[code] = parse_platform_csv(csv_bytes, code)

    if unknown_sheets:
        results["__unknown_sheets__"] = {"sheets": unknown_sheets}
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Parsing helpers
# ─────────────────────────────────────────────────────────────────────────────


def _normalise(s: str) -> str:
    """Lowercase + collapse separators so 'Amount_Spent (GBP)' matches 'amount spent'."""
    return (
        str(s or "")
        .strip()
        .lower()
        .replace("_", " ")
        .replace("-", " ")
        .replace("(gbp)", "")
        .replace("(usd)", "")
        .replace("(eur)", "")
        .replace("£", "")
        .strip()
    )


def _parse_number(raw: Any) -> Optional[float]:
    """Parse a single cell into a float; handle %, thousand separators, blanks."""
    if raw is None:
        return None
    s = str(raw).strip().replace(",", "").replace("£", "").replace("$", "").replace("€", "")
    if not s or s.lower() in ("--", "n/a", "na", "none"):
        return None
    is_percent = s.endswith("%")
    if is_percent:
        s = s[:-1].strip()
    try:
        v = float(s)
    except ValueError:
        return None
    if is_percent:
        v = v / 100.0
    return v


def _aggregate(rows: List[Dict[str, Any]], col: str, op: str) -> Optional[float]:
    """Reduce a column to a scalar.  op ∈ {'sum', 'mean', 'first', 'max'}."""
    vals: List[float] = []
    for row in rows:
        v = _parse_number(row.get(col))
        if v is not None:
            vals.append(v)
    if not vals:
        return None
    if op == "mean":
        return sum(vals) / len(vals)
    if op == "first":
        return vals[0]
    if op == "max":
        return max(vals)
    return sum(vals)


_TOTALS_ROW_MARKERS = ("total", "subtotal", "all conversions", "grand total")


def _is_totals_row(row: Dict[str, Any]) -> bool:
    """Detect summary rows that platforms append at the bottom of exports."""
    for v in row.values():
        if v is None:
            continue
        s = str(v).strip().lower()
        if not s:
            continue
        for marker in _TOTALS_ROW_MARKERS:
            if s.startswith(marker):
                return True
        return False
    return False


def _find_column(needles: Tuple[str, ...], column_index: Dict[str, str]) -> Optional[str]:
    """Return the actual CSV column name for the first matching needle."""
    for needle in needles:
        n = _normalise(needle)
        for norm_col, real_col in column_index.items():
            if n in norm_col:
                return real_col
    return None


def _aggregate_components_across_rows(
    rows: List[Dict[str, Any]],
    composition: KPIComposition,
    column_index: Dict[str, str],
) -> Tuple[Optional[float], List[Dict[str, Any]]]:
    """Aggregate one KPI from its composition rule, returning the composed
    value AND a per-component breakdown showing which columns matched.
    """
    component_values: List[Dict[str, Any]] = []
    for needle_group in composition.components:
        col = _find_column(needle_group, column_index)
        if col is None:
            continue
        # Aggregation per row-set depends on KPI type:
        #   - rates inside a composition operate on a single column via 'mean'
        #     across rows;
        #   - all other components sum within a column (one row per ad-set or
        #     campaign aggregates naturally).
        per_row_op = "mean" if composition.operator == "mean" else "sum"
        v = _aggregate(rows, col, per_row_op)
        if v is None:
            continue
        component_values.append({
            "needles": needle_group,
            "column": col,
            "value": v,
        })

    if not component_values:
        return None, []

    # Combine component values via the composition operator
    vals = [c["value"] for c in component_values]
    if composition.operator == "sum":
        total = sum(vals)
    elif composition.operator == "max":
        total = max(vals)
    elif composition.operator == "mean":
        total = sum(vals) / len(vals)
    else:  # "first"
        total = vals[0]

    return total, component_values


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────


def parse_platform_csv(
    content: Union[bytes, str],
    platform: str,
) -> Dict[str, Any]:
    """Parse a CSV export for one platform.

    Returns a dict shaped for Module 3 plus auditable composition data:
        {
          "budget": <float>,
          "historical_days": <int or None>,
          "kpis": {VAR: composed_value, ...},
          "kpi_breakdown": {
            VAR: {
              "value": composed_value,
              "operator": "sum" | "first" | ...,
              "rationale": "...",
              "components": [{column, value, needles}, ...],
            },
            ...
          },
          "matched_columns": {VAR: column_name, ...},
          "missing_kpis": [VAR, ...],
          "row_count": <int>,
        }

    Returns {"error": "..."} on a parse failure.
    """
    plat = (platform or "").strip().lower()
    if plat not in _CSV_PATTERNS:
        return {"error": f"CSV import is not supported for platform {plat!r}. "
                f"Supported: {', '.join(SUPPORTED_PLATFORMS)}."}

    if isinstance(content, bytes):
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                content = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return {"error": "Could not decode CSV — try saving as UTF-8."}

    try:
        sample = content[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(io.StringIO(content), dialect=dialect)
        rows = [r for r in reader if any((v or "").strip() for v in r.values())]
    except (csv.Error, ValueError) as e:
        return {"error": f"Could not parse CSV: {e}"}

    if not rows:
        return {"error": "CSV is empty or has no data rows."}

    rows = [r for r in rows if not _is_totals_row(r)]
    if not rows:
        return {"error": "CSV has no data rows after filtering totals."}

    column_index: Dict[str, str] = {}
    for col in rows[0].keys():
        if col is None:
            continue
        column_index.setdefault(_normalise(col), col)

    patterns = _CSV_PATTERNS[plat]
    kpis: Dict[str, float] = {}
    breakdown: Dict[str, Dict[str, Any]] = {}
    matched: Dict[str, str] = {}
    missing: List[str] = []
    budget_val: Optional[float] = None
    days_val: Optional[int] = None

    for var, comp in patterns.items():
        value, components = _aggregate_components_across_rows(rows, comp, column_index)
        # When primary composition matched nothing, try the documented
        # fallback (e.g. Meta's bundled 'Post engagement' when reactions
        # etc weren't exported separately).  The breakdown records which
        # composition won so the UI can show the fallback's warning.
        used_fallback = False
        if value is None and comp.fallback is not None:
            value, components = _aggregate_components_across_rows(
                rows, comp.fallback, column_index,
            )
            if value is not None:
                used_fallback = True
                comp = comp.fallback  # use fallback's rationale + operator

        if var == _BUDGET:
            budget_val = value
            if components:
                matched["budget"] = components[0]["column"]
            continue

        if var == _DAYS:
            if value is not None:
                days_val = int(value)
                matched["historical_days"] = components[0]["column"]
            continue

        if value is None or value <= 0:
            missing.append(var)
            continue

        # Rate KPIs live in [0, 1].  Some exports report percentages without
        # the '%' suffix (e.g. CTR shown as 4.50 meaning 4.5%); _parse_number
        # can't tell those apart from a literal 4.5.  If the composed rate
        # exceeds 1.0, treat the column as a bare-percentage form.
        if var in _RATE_KPIS and value > 1.0:
            if value <= 100.0:
                value = value / 100.0
                # Mirror the same correction in the component values for
                # auditability — otherwise the breakdown would look wrong.
                for c in components:
                    if c.get("value", 0.0) > 1.0:
                        c["value"] = c["value"] / 100.0
            else:
                missing.append(var)
                continue

        kpis[var] = value
        # First component's column is what we record for the legacy
        # matched_columns dict (back-compat with existing UI code).
        if components:
            matched[var] = components[0]["column"]
        breakdown[var] = {
            "value": value,
            "operator": comp.operator,
            "rationale": comp.rationale,
            "components": components,
            "used_fallback": used_fallback,
        }

    # Legacy rate fallback: if engagement is missing but the upload has a
    # rate column (e.g. an older export with 'Engagement Rate' but no
    # individual count breakouts) and an awareness count, derive
    # engagement = rate × awareness_count.  Keeps backward compatibility
    # with legacy exports without re-introducing rate units into the
    # canonical KPI set.
    if plat in _LEGACY_RATE_FALLBACK:
        rate_needles, awareness_var, target_var = _LEGACY_RATE_FALLBACK[plat]
        if target_var in missing and awareness_var in kpis:
            rate_value = None
            rate_column = None
            for needle in rate_needles:
                # Reuse the existing column-lookup machinery
                for norm_col, original_col in column_index.items():
                    if needle in norm_col:
                        agg = _aggregate(rows, original_col, "mean")
                        if agg is not None and agg > 0:
                            rate_value = agg
                            rate_column = original_col
                            break
                if rate_value is not None:
                    break
            if rate_value is not None:
                if rate_value > 1.0 and rate_value <= 100.0:
                    rate_value = rate_value / 100.0
                derived = rate_value * kpis[awareness_var]
                kpis[target_var] = derived
                missing.remove(target_var)
                matched[target_var] = rate_column
                breakdown[target_var] = {
                    "value": derived,
                    "operator": "rate_times_awareness",
                    "rationale": (
                        f"Engagement count derived from {rate_column} "
                        f"× {awareness_var} ({rate_value:.4f} × "
                        f"{kpis[awareness_var]:.0f}).  Fill the individual "
                        f"engagement count columns (Likes, Comments, etc.) "
                        f"to replace this with an exact sum."
                    ),
                    "components": [{
                        "column": rate_column, "value": rate_value,
                        "needle": "engagement rate (legacy)",
                    }],
                    "used_fallback": True,
                }

    return {
        "budget": budget_val,
        "historical_days": days_val,
        "kpis": kpis,
        "kpi_breakdown": breakdown,
        "matched_columns": matched,
        "missing_kpis": missing,
        "row_count": len(rows),
    }
